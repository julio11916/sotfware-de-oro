from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('bd/producto.xlsx')

# Get the active sheet
ws = wb.active

# Determine the last column
last_col = ws.max_column + 1

# Add the header 'eliminado' in the new column
ws.cell(row=1, column=last_col).value = 'eliminado'

# Set 'False' for all existing rows in the new column
for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=last_col).value = False

# Save the workbook
wb.save('bd/producto.xlsx')

print("Column 'eliminado' added successfully.")
