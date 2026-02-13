from openpyxl import load_workbook

# Load the workbook
wb = load_workbook('bd/producto.xlsx')
ws = wb.active

# Find the column index for 'eliminado'
eliminado_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == 'eliminado':
        eliminado_col = col
        break

if eliminado_col is None:
    print("Column 'eliminado' not found.")
    exit()

# Iterate through rows starting from 2 (skip header)
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=eliminado_col)
    val = cell.value
    if val is None or val == '' or str(val).lower() in ['nan', 'none']:
        cell.value = False
    else:
        # Convert to boolean
        cell.value = bool(val)

# Save the workbook
wb.save('bd/producto.xlsx')
print("Column 'eliminado' fixed: all values are now boolean (True/False).")
