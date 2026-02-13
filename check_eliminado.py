from openpyxl import load_workbook

wb = load_workbook('bd/producto.xlsx')
ws = wb.active

# Find column
col = None
for c in range(1, ws.max_column + 1):
    if ws.cell(1, c).value == 'eliminado':
        col = c
        break

print(f"Columna 'eliminado': {col}")
print("Valores en la columna:")
for r in range(1, ws.max_row + 1):
    val = ws.cell(r, col).value
    print(f"Fila {r}: {val} (tipo: {type(val)})")
